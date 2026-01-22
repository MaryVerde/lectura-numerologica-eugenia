import os
import unicodedata
import re
from datetime import date,datetime
from io import BytesIO
import textwrap
import hmac
import hashlib

import streamlit as st
from reportlab.lib.pagesizes import LETTER
from reportlab.pdfgen import canvas
from collections import Counter
from openpyxl import load_workbook
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.colors import HexColor

if "premium_activo" not in st.session_state:
    st.session_state.premium_activo = False



# =====================================================
# SECRETOS (STREAMLIT CLOUD + LOCAL)
# =====================================================
def get_secret(key: str, default=None):
    # 1) Streamlit Secrets
    try:
        if hasattr(st, "secrets") and key in st.secrets:
            return st.secrets[key]
    except Exception:
        pass
    # 2) Variables de entorno (local)
    return os.getenv(key, default)

APP_SECRET = get_secret("APP_SECRET")
ADMIN_PIN = get_secret("ADMIN_PIN")

if not APP_SECRET:
    st.error("❌ Falta APP_SECRET. Ve a Settings → Secrets y agrega APP_SECRET.")
    st.stop()

# =====================================================
# CONTADOR (INTERNO) - SOLO PANEL ADMIN
# =====================================================
COUNTER_FILE = "contador_resumida.txt"

def leer_contador():
    try:
        with open(COUNTER_FILE, "r", encoding="utf-8") as f:
            return int(f.read().strip())
    except:
        return 0

def incrementar_contador():
    total = leer_contador() + 1
    try:
        with open(COUNTER_FILE, "w", encoding="utf-8") as f:
            f.write(str(total))
    except:
        # En Streamlit Cloud a veces el FS es de solo lectura
        pass
    return total

# ==============================================
# CONFIGURACIÓN GENERAL
# ==============================================

APP_TITLE = "🔮 Lectura Numerológica"
BRAND = "Eugenia.Mystikos"

st.set_page_config(
    page_title=f"{APP_TITLE} · {BRAND}",
    page_icon="🔮",
    layout="centered"
)

# --- ESTILO VISUAL (marca en rojo) ---
st.markdown("""
<style>
h1 {
    color: #b11226;
    font-weight: 700;
}
.brand {
    color: #b11226;
    font-weight: 600;
}
.subtitle {
    color: #444444;
    font-size: 0.95rem;
}
</style>
""", unsafe_allow_html=True)

# --- TÍTULO ---
st.markdown(
    "<h1>🔮 Lectura Numerológica · <span class='brand'>Eugenia.Mystikos</span></h1>",
    unsafe_allow_html=True
)

st.markdown(
    "<div class='subtitle'>"
    "Versión Resumida · Interpretación completa disponible en versión Premium (PDF personalizado)"
    "</div>",
    unsafe_allow_html=True
)
# =====================================================
# UTILIDADES NUMEROLÓGICAS
# =====================================================
MASTER = {11, 22, 33}

def reducir_numero(n: int) -> int:
    n = abs(int(n))
    if n in MASTER:
        return n
    while n > 9:
        n = sum(int(d) for d in str(n))
        if n in MASTER:
            return n
    return n

def normalizar_texto(s: str) -> str:
    s = unicodedata.normalize("NFD", str(s))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s.upper()

TABLA_PITAGORICA = {
    **{c: 1 for c in "AJS"},
    **{c: 2 for c in "BKT"},
    **{c: 3 for c in "CLU"},
    **{c: 4 for c in "DMV"},
    **{c: 5 for c in "ENW"},
    **{c: 6 for c in "FOX"},
    **{c: 7 for c in "GPY"},
    **{c: 8 for c in "HQZ"},
    **{c: 9 for c in "IR"},
}

def numero_nombre(nombre: str) -> int:
    total = 0
    for c in normalizar_texto(nombre):
        if c.isalpha():
            total += TABLA_PITAGORICA.get(c, 0)
    return reducir_numero(total)

def sumar_digitos_texto(txt: str) -> int:
    digs = re.findall(r"\d", str(txt))
    if not digs:
        return 0
    return reducir_numero(sum(int(d) for d in digs))

def numero_apto(apto: str) -> int:
    apto = str(apto).strip()
    if not apto:
        return 0
    if re.search(r"\d", apto):
        return sumar_digitos_texto(apto)
    return numero_nombre(apto)

# ---- Núcleos principales ----
def esencia(fecha: date) -> int:
    return reducir_numero(fecha.day)

def vida_pasada(fecha: date) -> int:
    return reducir_numero(fecha.month)

def sendero_vida(fecha: date) -> int:
    return reducir_numero(fecha.day + fecha.month + fecha.year)

def ano_personal(fecha: date, year: int) -> int:
    return reducir_numero(fecha.day + fecha.month + year)

def mes_personal(ano_p: int, mes: int) -> int:
    return reducir_numero(ano_p + mes)

def semana_personal(mes_p: int, semana_del_ano: int) -> int:
    return reducir_numero(mes_p + semana_del_ano)

def dia_personal(mes_p: int, dia_hoy: int) -> int:
    return reducir_numero(mes_p + dia_hoy)

# ---- Arcano semanal ----
def arcano_semanal() -> int:
    semana = date.today().isocalendar()[1]
    return (semana % 22) + 1

# ---- Pináculo pirámide completa ----
def pinaculo_piramide(fecha: date) -> dict:
    d = reducir_numero(fecha.day)
    m = reducir_numero(fecha.month)
    a = reducir_numero(fecha.year)

    p1 = reducir_numero(d + m)
    p2 = reducir_numero(d + a)
    p3 = reducir_numero(p1 + p2)

    p4 = reducir_numero(p1 + p2)
    p5 = reducir_numero(p2 + p3)

    p6 = reducir_numero(p4 + p5)

    return {"base": (p1, p2, p3), "medio": (p4, p5), "cima": p6}

# =====================================================
# TEXTOS RESUMIDOS (base)
# =====================================================
LECTURA_RESUMIDA = {
    1:  "Te invita a marca un renacer personal. La vida te coloca frente a decisiones que no pueden seguir postergándose. Se activa el fuego del inicio, la valentía de decir “sí” a lo nuevo y “no” a lo que ya no vibra contigo. Todo te empuja a tomar liderazgo sobre tu propia historia. No esperes señales externas: la señal eres tú. Lo que comiences ahora define el tono de los próximos años. Este es un año para actuar con claridad, coraje y propósito. La energía te respalda cuando confías en tu impulso interior.",
    2:  "Te invita a afinar la sensibilidad y profundizar los vínculos. La vida te enseña que no todo se logra empujando: algunas cosas florecen cuando aprendes a escuchar. Se activa la energía de la cooperación, la paciencia y la armonía. Es un ciclo para sanar relaciones, equilibrar emociones y reconocer que la verdadera fortaleza también sabe esperar. El crecimiento llega cuando honras los ritmos naturales y eliges la paz sin perderte a ti.",
    3:  "Te invita a despierta tu voz auténtica y tu creatividad. La energía te empuja a expresarte, a mostrarte y a disfrutar más del proceso de vivir. Se abre un ciclo donde la alegría no es superficial, sino medicina. Todo lo que comunicas tiene impacto, por eso es importante hablar desde la verdad. Es un año para crear, compartir, conectar y permitir que tu luz sea vista. Cuando te permites ser tú, la vida responde con expansión.",
    4:  "Te invita a tener orden, estructura y compromiso contigo misma. No es un ciclo de velocidad, sino de construcción consciente. La energía te invita a poner bases sólidas para el futuro, incluso si eso requiere disciplina y constancia. Cada paso cuenta, aunque no lo veas de inmediato. Es un año para materializar con paciencia, organizar prioridades y fortalecer lo que realmente importa. Lo que edificas ahora tiene raíces profundas.",
    5:  "Te invita a trae cambio, movimiento y liberación. La vida sacude lo que estaba estancado y te invita a salir de lo conocido. Se activa una energía inquieta que pide experiencias nuevas, decisiones valientes y flexibilidad. Resistirse solo genera tensión: fluir abre caminos inesperados. Es un año para reinventarte, viajar interna o externamente, y recordar que la libertad también es una elección consciente.",
    6:  "Te invita a poner foco está en el corazón, el cuidado y la responsabilidad emocional. La energía te lleva a revisar vínculos, compromisos y la forma en que das y recibes amor. Es un ciclo de sanación afectiva, donde se te pide equilibrio entre cuidar a otros y cuidarte a ti. El hogar interno se vuelve prioridad. Cuando eliges desde el amor consciente, todo se ordena con mayor armonía.",
    7:  "Te invita a hacer un viaje interior profundo. La vida baja el ruido externo para que puedas escuchar tu sabiduría interna. Se activa la introspección, la búsqueda de sentido y la conexión espiritual. No es un año para forzar resultados, sino para comprender procesos. El silencio se vuelve aliado. Las respuestas llegan cuando confías en tu intuición y honras tus tiempos internos.",
    8:  "Te invita a activa el poder personal, la autoridad interna y la manifestación. La energía te confronta con temas de merecimiento, límites y abundancia. Es un ciclo para tomar control consciente de tu vida material y emocional. El éxito llega cuando actúas con integridad y coherencia. Es un año para asumir tu fuerza sin culpa y reconocer el valor real de lo que aportas al mundo.",
    9:  "Te invita a marca un cierre de ciclo profundo. La vida te invita a soltar lo que ya cumplió su función: relaciones, patrones, historias y cargas emocionales. Es un año de limpieza, perdón y liberación. No se trata de pérdida, sino de preparación para un nuevo comienzo. Al dejar ir, recuperas energía vital. La sabiduría adquirida es tu mayor tesoro.",
    11: "Te invita a despierta una conciencia elevada y una sensibilidad espiritual intensa. La energía te convierte en canal de inspiración, intuición y guía. Puedes sentir todo más fuerte, pero también ver más claro. Es un año para confiar en tu percepción, cuidar tu energía y honrar tu luz. Cuando te alineas con tu verdad, impactas más de lo que imaginas.",
    22: "Te invita a activar la energía del gran constructor. La visión se une a la acción y te pide materializar algo con propósito colectivo. No es un ciclo liviano: implica responsabilidad, compromiso y enfoque. Pero también ofrece la posibilidad de crear algo duradero y significativo. Cuando alineas intención y acción, puedes dejar huella real en el mundo.",
    33: "Te invita a orientar al amor consciente y al servicio con madurez emocional. Invita a acompañar sin rescatar y a dar sin vaciarte. Tu sensibilidad se vuelve fortaleza cuando hay límites, estructura y autocuidado.",
}

def lectura_resumida(num: int) -> str:
    return LECTURA_RESUMIDA.get(num, "Lectura no disponible para esta vibración.")

# =====================================================
# GRATIS: FRASES CORTAS (AMOR / DINERO / EMOCIONAL / PROTECCIÓN)
# Basadas en tu Año Personal (ap)
# =====================================================
FRASES_AMOR = {
    1:"Amor: inicia desde ti; el vínculo correcto nace cuando eliges con valentía y dejas de mendigar señales.",
    2:"Amor: escucha y suaviza; lo que crece en silencio se vuelve sólido cuando hay respeto y paciencia.",
    3:"Amor: habla claro; tu encanto abre puertas, pero tu verdad sostiene lo que merece quedarse.",
    4:"Amor: construye con hechos; promesas sin estructura se caen, límites sanos se quedan.",
    5:"Amor: cambia la dinámica; si te sientes atrapada, es hora de reinventar la manera de amar.",
    6:"Amor: cuida sin cargarte; equilibrio entre dar y recibir es la medicina del vínculo.",
    7:"Amor: baja el ruido; la intuición muestra quién suma paz y quién consume energía.",
    8:"Amor: merecimiento; el vínculo se ordena cuando tú te valoras y sostienes tu lugar.",
    9:"Amor: cierre limpio; lo que termina te libera para amar con más conciencia.",
    11:"Amor: sensibilidad elevada; protege tu energía, elige vínculos que honren tu luz.",
    22:"Amor: proyecto en común; el vínculo crece cuando hay visión, madurez y acuerdos reales.",
    33:"Amor: amor consciente; acompaña sin salvar y ama sin vaciarte."
}
FRASES_DINERO = {
    1:"Dinero: actúa y decide; este año premia el liderazgo y castiga la duda eterna.",
    2:"Dinero: alianza y paciencia; creces más si negocias con calma y construyes relaciones.",
    3:"Dinero: visibilidad; comunicar y mostrar tu talento abre oportunidades y expansión.",
    4:"Dinero: estructura; presupuesto, orden y disciplina convierten esfuerzo en estabilidad.",
    5:"Dinero: movimiento; diversifica, prueba, adapta: la rigidez aquí se rompe.",
    6:"Dinero: responsabilidad; prosperas cuando cuidas compromisos y pones precio a tu entrega.",
    7:"Dinero: estrategia; menos impulso, más análisis: invertir en conocimiento rinde.",
    8:"Dinero: poder y abundancia; liderazgo con ética = resultados reales.",
    9:"Dinero: cierre y depuración; suelta fugas y deudas emocionales para liberar flujo.",
    11:"Dinero: inspiración con enfoque; baja ideas a plan y sostén tu energía.",
    22:"Dinero: construcción grande; visión + método = legado material sostenible.",
    33:"Dinero: servicio consciente; prosperas cuando tu aporte transforma y tiene límites."
}
FRASES_EMOCIONAL = {
    1:"Emocional: reafirma tu voz; no te traiciones por encajar.",
    2:"Emocional: regula y escucha; tu calma es tu superpoder.",
    3:"Emocional: expresa sin drama; lo que nombras se ordena.",
    4:"Emocional: estabilidad; rutina y límites te devuelven centro.",
    5:"Emocional: libertad; el cambio es medicina si lo eliges con conciencia.",
    6:"Emocional: corazón; aprende a cuidar sin cargarte.",
    7:"Emocional: introspección; tu alma pide silencio y claridad.",
    8:"Emocional: fuerza; no confundas control con seguridad: elige coherencia.",
    9:"Emocional: cierre; perdonar es liberar energía, no justificar.",
    11:"Emocional: sensibilidad; filtra ambientes y respira antes de decidir.",
    22:"Emocional: responsabilidad; madurez afectiva para sostener lo grande.",
    33:"Emocional: compasión; amor con límites para no agotarte."
}
FRASES_PROTECCION = {
    1:"Protección: corta lo tibio; tu energía se protege cuando dices ‘no’ sin culpa.",
    2:"Protección: límites suaves; no todo merece acceso a tu intimidad.",
    3:"Protección: palabra consciente; evita prometer desde emoción, elige claridad.",
    4:"Protección: orden y tierra; tu rutina es tu escudo energético.",
    5:"Protección: evita excesos; libertad sí, caos no.",
    6:"Protección: hogar interno; cuida tu descanso, tu cuerpo y tus vínculos.",
    7:"Protección: silencio; menos exposición, más intuición.",
    8:"Protección: autoridad; protege tu valor y tu tiempo como oro.",
    9:"Protección: limpieza; suelta culpas, cierra puertas con dignidad.",
    11:"Protección: alta vibración; filtra personas y ambientes, elige lo sagrado.",
    22:"Protección: enfoque; grandes metas requieren límites firmes.",
    33:"Protección: amor consciente; dar con estructura, no desde sacrificio."
}

def frase_categoria(dic: dict, num: int) -> str:
    return dic.get(num, "Mensaje no disponible para esta vibración.")

# =====================================================
# # 🌅 ENERGÍA DEL DÍA (365 mensajes) — REGALO (EXPRESS)
# =====================================================
ENERGIA_DIA_365 = {
    1: "HOY NO APRESURES NADA. LA ENERGÍA SE ORDENA CUANDO ELIGES PRESENCIA EN LUGAR DE URGENCIA.",
    2: "CONFÍA EN TU RITMO. NO TODO FLORECE EL MISMO DÍA, PERO TODO RESPONDE A LA INTENCIÓN CORRECTA.",
    3: "LO QUE HOY PARECE PEQUEÑO ESTÁ SEMBRANDO UNA VERDAD MÁS GRANDE.",
    4: "RESPIRA ANTES DE DECIDIR. LA CLARIDAD LLEGA CUANDO EL CUERPO SE RELAJA.",
    5: "NO TE ADAPTES A LO QUE TE APAGA. AJUSTA EL ENTORNO, NO TU ESENCIA.",
    6: "HOY ES UN BUEN DÍA PARA PONER UN LÍMITE AMOROSO.",
    7: "EL SILENCIO TAMBIÉN ES UNA RESPUESTA SABIA.",
    8: "SUELTA EL CONTROL: LO VERDADERO NO NECESITA SER FORZADO.",
    9: "HOY HONRA LO QUE YA LOGRASTE. RECONOCER TU AVANCE CAMBIA LA ENERGÍA.",
    10: "LA COHERENCIA VALE MÁS QUE LA VELOCIDAD.",
    11: "TU SENSIBILIDAD ES UNA BRÚJULA, NO UNA DEBILIDAD.",
    12: "ESCUCHA LO QUE INCOMODA: AHÍ HAY INFORMACIÓN VALIOSA.",
    13: "CERRAR A TIEMPO TAMBIÉN ES UN ACTO DE AMOR PROPIO.",
    14: "HOY ELIGE CON CALMA, INCLUSO SI OTROS APURAN.",
    15: "NO TODO MERECE TU ENERGÍA. SÉ SELECTIVA.",
    16: "LA VERDAD SE SOSTIENE SOLA. NO LA JUSTIFIQUES.",
    17: "HOY EL CUERPO SABE MÁS QUE LA MENTE.",
    18: "AVANZA UN PASO REAL, NO DIEZ IMAGINARIOS.",
    19: "TU INTUICIÓN ESTÁ CLARA CUANDO NO LA DISCUTES.",
    20: "ORDEN EXTERNO, PAZ INTERNA.",
    21: "HOY SE AFLOJA UNA CARGA QUE NO ERA TUYA.",
    22: "CONFÍA: LO QUE SE ACOMODA HOY LIBERA FUTURO.",
    23: "NO TE TRAICIONES PARA EVITAR CONFLICTO.",
    24: "LA ENERGÍA RESPONDE A LA HONESTIDAD.",
    25: "DESCANSAR TAMBIÉN ES AVANZAR.",
    26: "HOY ELIGE LO SIMPLE. AHÍ ESTÁ LA FUERZA.",
    27: "NO RESCATES PROCESOS AJENOS.",
    28: "TU CLARIDAD INSPIRA SIN QUE HABLES.",
    29: "HOY ES MEJOR DECIR MENOS Y SENTIR MÁS.",
    30: "LA ESTABILIDAD SE CONSTRUYE CON DECISIONES PEQUEÑAS.",
    31: "CIERRA EL MES SOLTANDO EXPECTATIVAS IRREALES.",
    32: "HOY TU ENERGÍA PIDE ENFOQUE, NO DISPERSIÓN.",
    33: "ELEGIR PAZ NO ES RENDIRSE.",
    34: "NO NEGOCIES LO ESENCIAL.",
    35: "LA VIDA TE RESPONDE CUANDO TE ALINEAS.",
    36: "HOY SE ORDENA ALGO INTERNO SI NO LO FUERZAS.",
    37: "OBSERVA SIN JUZGAR: AHÍ ESTÁ LA ENSEÑANZA.",
    38: "NO TODO SE RESUELVE HOY, Y ESTÁ BIEN.",
    39: "RESPETA TU PROCESO AUNQUE OTROS NO LO ENTIENDAN.",
    40: "TU ENERGÍA VALE MÁS QUE TU EXPLICACIÓN.",
    41: "HOY ES DÍA DE SOSTENER, NO DE EMPUJAR.",
    42: "CUANDO TE ELIGES, TODO SE REACOMODA.",
    43: "NO RESPONDAS DESDE LA HERIDA.",
    44: "EL EQUILIBRIO SE CONSTRUYE CON LÍMITES CLAROS.",
    45: "HOY TU PRESENCIA ES SUFICIENTE.",
    46: "LA CALMA TAMBIÉN ES PODER.",
    47: "NO CORRIJAS LO QUE AÚN ESTÁ APRENDIENDO.",
    48: "HOY ESCUCHA TU CANSANCIO CON RESPETO.",
    49: "LO QUE SE VA LIBERA ESPACIO.",
    50: "AVANZA SIN RUIDO, PERO CON CERTEZA.",
    51: "NO PROMETAS DESDE LA EMOCIÓN.",
    52: "EL CUERPO PIDE VERDAD, NO DISCURSO.",
    53: "HOY CUIDA TU ENERGÍA COMO ALGO SAGRADO.",
    54: "NO TODO MERECE RESPUESTA INMEDIATA.",
    55: "ELEGIR DISTINTO ES EVOLUCIÓN.",
    56: "LA CLARIDAD LLEGA CUANDO DEJAS DE JUSTIFICAR.",
    57: "HOY HONRA TUS LÍMITES.",
    58: "NO CARGUES CON LO QUE NO TE CORRESPONDE.",
    59: "LA COHERENCIA SE SIENTE.",
    60: "SUELTA LA EXPECTATIVA, SOSTÉN LA INTENCIÓN.",
    61: "HOY EL ORDEN INTERNO ES PRIORIDAD.",
    62: "TU ENERGÍA SE EXPANDE CUANDO TE RESPETAS.",
    63: "NO EXPLIQUES TU VERDAD: VÍVELA.",
    64: "HOY ES MEJOR AVANZAR LENTO QUE DUDAR RÁPIDO.",
    65: "LA ESTABILIDAD NACE DE DECISIONES HONESTAS.",
    66: "NO TE ADAPTES A LO QUE TE DRENA.",
    67: "LA VIDA RESPONDE A TU CLARIDAD.",
    68: "HOY ESCUCHA SIN INTERRUMPIRTE.",
    69: "EL SILENCIO ORDENA MÁS DE LO QUE CREES.",
    70: "TU INTUICIÓN ESTÁ AFINADA.",
    71: "NO TODO CIERRE ES PÉRDIDA.",
    72: "HOY SUELTA LA AUTOEXIGENCIA INNECESARIA.",
    73: "RESPETA TUS TIEMPOS INTERNOS.",
    74: "ELEGIR CALMA ES ELEGIR PODER.",
    75: "NO TE DISTRAIGAS DE LO IMPORTANTE.",
    76: "HOY CUIDA TU ENERGÍA EMOCIONAL.",
    77: "LA CLARIDAD NO GRITA.",
    78: "NO RESCATES PROCESOS QUE NO SON TUYOS.",
    79: "TU PAZ ES PRIORIDAD.",
    80: "HOY SE ORDENA ALGO SI NO INTERVIENES DE MÁS.",
    81: "AVANZA CON FIRMEZA TRANQUILA.",
    82: "NO FUERCES ACUERDOS.",
    83: "EL EQUILIBRIO SE CONSTRUYE.",
    84: "HOY ESCUCHA TU CUERPO.",
    85: "NO TODO SE DECIDE HOY.",
    86: "LA COHERENCIA TE SOSTIENE.",
    87: "SUELTA LO QUE PESA.",
    88: "HOY CONFÍA EN LO QUE SIENTES.",
    89: "NO TE JUSTIFIQUES.",
    90: "LA ENERGÍA RESPONDE A TU HONESTIDAD.",
    91: "HOY ELIGE PRESENCIA ANTES QUE REACCIÓN.",
    92: "LA CLARIDAD SE ACTIVA CUANDO DEJAS DE FORZAR.",
    93: "HOY CONFÍA EN LO QUE YA SABES INTERNAMENTE.",
    94: "NO TODO REQUIERE RESPUESTA INMEDIATA.",
    95: "TU ENERGÍA SE ORDENA CUANDO TE RESPETAS.",
    96: "HOY MENOS PALABRAS, MÁS VERDAD.",
    97: "EL EQUILIBRIO NACE DE DECISIONES PEQUEÑAS.",
    98: "HOY TU CUERPO HABLA: ESCÚCHALO.",
    99: "LA CALMA TAMBIÉN ES ACCIÓN.",
    100: "HOY SOSTÉN TU CENTRO SIN EXPLICARTE.",
    101: "NO TE DISPERSES: VUELVE A LO ESENCIAL.",
    102: "HOY SUELTA LA PRISA, NO EL RUMBO.",
    103: "ELEGIR PAZ ES UN ACTO DE PODER.",
    104: "HOY HONRA TUS LÍMITES.",
    105: "LO ALINEADO NO SE SIENTE PESADO.",
    106: "RESPIRA ANTES DE DECIDIR.",
    107: "NO CARGUES LO QUE NO TE CORRESPONDE.",
    108: "HOY LA COHERENCIA ES PROTECCIÓN.",
    109: "AVANZA SIN JUSTIFICARTE.",
    110: "TU ENERGÍA RESPONDE A TU HONESTIDAD.",
    111: "HOY TU INTUICIÓN ESTÁ AFINADA.",
    112: "NO FUERCES ACUERDOS.",
    113: "EL ORDEN INTERNO SE REFLEJA AFUERA.",
    114: "HOY ELIGE CALIDAD, NO CANTIDAD.",
    115: "SUELTA EL CONTROL EXCESIVO.",
    116: "LO SIMPLE TAMBIÉN ES SAGRADO.",
    117: "HOY CUIDA TU ENERGÍA EMOCIONAL.",
    118: "NO TODO SE DECIDE HOY.",
    119: "ESCUCHA MÁS DE LO QUE HABLAS.",
    120: "TU PRESENCIA ES SUFICIENTE.",
    121: "HOY EL SILENCIO TRAE CLARIDAD.",
    122: "NO TE TRAICIONES POR COMODIDAD.",
    123: "EL DESCANSO TAMBIÉN ES PRODUCTIVIDAD.",
    124: "HOY AVANZA SIN RUIDO.",
    125: "CONFÍA EN EL PROCESO QUE YA EMPEZÓ.",
    126: "TU CENTRO ES TU GUÍA.",
    127: "NO EXPLIQUES LO QUE YA SENTISTE.",
    128: "HOY BAJA EL RITMO CONSCIENTEMENTE.",
    129: "LO VERDADERO NO SE APURA.",
    130: "TU PAZ ES PRIORIDAD.",
    131: "HOY OBSERVA ANTES DE ACTUAR.",
    132: "NO TODO REQUIERE INTERVENCIÓN.",
    133: "LA CLARIDAD LLEGA CUANDO PARAS.",
    134: "HOY CUIDA TUS PALABRAS.",
    135: "NO CARGUES EXPECTATIVAS AJENAS.",
    136: "EL EQUILIBRIO SE CONSTRUYE.",
    137: "HOY ELIGE PRESENCIA CORPORAL.",
    138: "LA CALMA ORDENA DECISIONES.",
    139: "SUELTA LA NECESIDAD DE CONVENCER.",
    140: "TU ENERGÍA SE REAJUSTA SOLA.",
    141: "HOY VUELVE A LO ESENCIAL.",
    142: "NO TE DISPERSES EMOCIONALMENTE.",
    143: "EL FOCO ES MEDICINA.",
    144: "HOY HONRA TU RITMO INTERNO.",
    145: "NO TODO MERECE RESPUESTA.",
    146: "TU COHERENCIA ABRE CAMINO.",
    147: "LA CLARIDAD NO GRITA.",
    148: "HOY CONFÍA SIN FORZAR.",
    149: "SOSTÉN TU VERDAD CON CALMA.",
    150: "MENOS RUIDO, MÁS CENTRO.",
    151: "HOY ELIGE ESTABILIDAD EMOCIONAL.",
    152: "NO REACCIONES DESDE EL CANSANCIO.",
    153: "EL ORDEN INTERNO SE NOTA.",
    154: "HOY NO TE SOBREEXIJAS.",
    155: "LA PAUSA ES PARTE DEL AVANCE.",
    156: "TU ENERGÍA SE REGULA CON LÍMITES.",
    157: "HOY RESPIRA CONSCIENTEMENTE.",
    158: "NO FUERCES RESULTADOS.",
    159: "EL CUERPO MARCA EL CAMINO.",
    160: "HOY SOSTÉN TU EJE.",
    161: "NO TE ADELANTES AL PROCESO.",
    162: "HOY ESCUCHA SIN DEFENDERTE.",
    163: "LA SERENIDAD ES PODER.",
    164: "NO TODO ES URGENTE.",
    165: "HOY ELIGE CLARIDAD INTERNA.",
    166: "SUELTA LA AUTOEXIGENCIA.",
    167: "LA CALMA TE ORDENA.",
    168: "HOY CUIDA TU ENERGÍA VITAL.",
    169: "NO CARGUES LO INNECESARIO.",
    170: "TU CENTRO TE SOSTIENE.",
    171: "HOY CONFÍA EN EL PASO PRESENTE.",
    172: "NO TODO SE RESUELVE HOY.",
    173: "LA COHERENCIA TE PROTEGE.",
    174: "HOY ELIGE SOBRIEDAD EMOCIONAL.",
    175: "NO TE PIERDAS POR COMPLACER.",
    176: "EL SILENCIO TAMBIÉN COMUNICA.",
    177: "HOY BAJA EXPECTATIVAS EXTERNAS.",
    178: "TU ENERGÍA SE AFINA SOLA.",
    179: "EL EQUILIBRIO ES PRÁCTICA DIARIA.",
    180: "HOY SOSTÉN LO QUE ES REAL.",
    181: "NO FUERCES CONVERSACIONES.",
    182: "HOY PRIORIZA TU ESTABILIDAD.",
    183: "LA CLARIDAD SE CONSTRUYE.",
    184: "NO TOMES DECISIONES CANSADA.",
    185: "HOY HONRA TU INTUICIÓN.",
    186: "LA CALMA ES DIRECCIÓN.",
    187: "NO TE EXPLIQUES DE MÁS.",
    188: "HOY ELIGE SENCILLEZ.",
    189: "TU ENERGÍA PIDE ORDEN.",
    190: "SUELTA LO QUE PESA.",
    191: "HOY VUELVE A TU CUERPO.",
    192: "NO PERSIGAS RESPUESTAS.",
    193: "LA PRESENCIA ES SUFICIENTE.",
    194: "HOY CUIDA TUS LÍMITES.",
    195: "NO CARGUES CULPAS AJENAS.",
    196: "EL CENTRO SE RECUPERA.",
    197: "HOY ACTÚA CON MESURA.",
    198: "LA CALMA ESTABILIZA.",
    199: "NO TODO SE EXPLICA.",
    200: "HOY ELIGE COHERENCIA.",
    201: "RESPETA TU ENERGÍA.",
    202: "NO TE FUERCES A RENDIR.",
    203: "LA CLARIDAD LLEGA SOLA.",
    204: "HOY BAJA EL RUIDO MENTAL.",
    205: "NO TE DISPERSES EMOCIONALMENTE.",
    206: "EL EQUILIBRIO ES INTERNO.",
    207: "HOY CONFÍA EN TU PROCESO.",
    208: "NO TODO SE COMPARTE.",
    209: "LA SOBRIEDAD PROTEGE.",
    210: "HOY VUELVE A TU EJE.",
    211: "NO TE ADELANTES.",
    212: "LA PAUSA ES SABIA.",
    213: "HOY ESCUCHA TU CUERPO.",
    214: "NO CARGUES TENSIONES VIEJAS.",
    215: "EL PRESENTE BASTA.",
    216: "HOY ELIGE CALMA.",
    217: "NO REACCIONES POR HÁBITO.",
    218: "TU ENERGÍA SE REGULA.",
    219: "LA CLARIDAD NO SE FUERZA.",
    220: "HOY SOSTÉN TU VERDAD.",
    221: "NO TE PIERDAS EN RUIDO EXTERNO.",
    222: "HOY CUIDA TU CENTRO.",
    223: "EL EQUILIBRIO SE SIENTE.",
    224: "NO TODO ES PRIORIDAD.",
    225: "HOY BAJA EL RITMO.",
    226: "LA CALMA ES ESTRATEGIA.",
    227: "NO TE DISPERSES.",
    228: "HOY RESPIRA PROFUNDO.",
    229: "LA COHERENCIA ORDENA.",
    230: "TU ENERGÍA RESPONDE.",
    231: "NO FUERCES SOLUCIONES.",
    232: "HOY ELIGE PRESENCIA.",
    233: "EL SILENCIO ACLARA.",
    234: "NO TODO SE RESUELVE HOY.",
    235: "HOY CONFÍA EN TU CENTRO.",
    236: "LA CALMA GUÍA.",
    237: "NO CARGUES EXPECTATIVAS.",
    238: "HOY SOSTÉN TU EJE.",
    239: "LA SOBRIEDAD ES FUERZA.",
    240: "TU ENERGÍA SE ORDENA.",
    241: "NO TE DISPERSES MENTALMENTE.",
    242: "HOY PRIORIZA LO ESENCIAL.",
    243: "EL EQUILIBRIO SE CONSTRUYE.",
    244: "NO FUERCES RITMOS.",
    245: "HOY ESCUCHA MÁS.",
    246: "LA PRESENCIA SANA.",
    247: "NO CARGUES TENSIONES.",
    248: "HOY ELIGE COHERENCIA.",
    249: "LA CALMA SOSTIENE.",
    250: "TU CENTRO ES GUÍA.",
    251: "NO TODO SE DECIDE HOY.",
    252: "HOY BAJA LA EXIGENCIA.",
    253: "EL SILENCIO PROTEGE.",
    254: "NO REACCIONES AUTOMÁTICAMENTE.",
    255: "HOY HONRA TU CUERPO.",
    256: "LA CLARIDAD LLEGA.",
    257: "NO FUERCES RESPUESTAS.",
    258: "HOY CONFÍA EN TI.",
    259: "EL EQUILIBRIO SE AFINA.",
    260: "TU ENERGÍA RESPONDE.",
    261: "NO CARGUES LO INNECESARIO.",
    262: "HOY ELIGE CALMA INTERNA.",
    263: "LA COHERENCIA GUÍA.",
    264: "NO TE DISPERSES.",
    265: "HOY RESPIRA PROFUNDO.",
    266: "LA SOBRIEDAD ORDENA.",
    267: "NO FUERCES PROCESOS.",
    268: "HOY SOSTÉN TU CENTRO.",
    269: "LA PRESENCIA BASTA.",
    270: "TU ENERGÍA SE ALINEA.",
    271: "NO TE ADELANTES.",
    272: "HOY CUIDA TU RITMO.",
    273: "EL SILENCIO ACLARA.",
    274: "NO CARGUES RUIDO.",
    275: "HOY ELIGE ESTABILIDAD.",
    276: "LA CALMA ES DIRECCIÓN.",
    277: "NO FUERCES ACUERDOS.",
    278: "HOY ESCUCHA TU CUERPO.",
    279: "EL EQUILIBRIO PROTEGE.",
    280: "TU CENTRO RESPONDE.",
    281: "NO REACCIONES POR COSTUMBRE.",
    282: "HOY BAJA EL RITMO.",
    283: "LA CLARIDAD SE SIENTE.",
    284: "NO TE SOBREEXIJAS.",
    285: "HOY HONRA TU ENERGÍA.",
    286: "LA COHERENCIA SOSTIENE.",
    287: "NO CARGUES TENSIONES.",
    288: "HOY ELIGE PRESENCIA.",
    289: "EL SILENCIO ORDENA.",
    290: "TU ENERGÍA RESPONDE.",
    291: "NO FUERCES RESULTADOS.",
    292: "HOY VUELVE A LO SIMPLE.",
    293: "LA CALMA GUÍA.",
    294: "NO TE DISPERSES.",
    295: "HOY ESCUCHA MÁS.",
    296: "EL EQUILIBRIO SE AJUSTA.",
    297: "NO CARGUES EXPECTATIVAS.",
    298: "HOY CONFÍA EN TU CENTRO.",
    299: "LA PRESENCIA BASTA.",
    300: "TU ENERGÍA SE ORDENA.",
    301: "NO TODO SE RESUELVE HOY.",
    302: "HOY BAJA LA PRISA.",
    303: "LA COHERENCIA PROTEGE.",
    304: "NO TE FUERCES.",
    305: "HOY HONRA TU RITMO.",
    306: "EL SILENCIO ACLARA.",
    307: "NO CARGUES RUIDO.",
    308: "HOY ELIGE CALMA.",
    309: "EL EQUILIBRIO SOSTIENE.",
    310: "TU CENTRO GUÍA.",
    311: "NO REACCIONES AUTOMÁTICAMENTE.",
    312: "HOY ESCUCHA TU CUERPO.",
    313: "LA CLARIDAD SE SIENTE.",
    314: "NO TE DISPERSES.",
    315: "HOY CUIDA TU ENERGÍA.",
    316: "LA COHERENCIA ORDENA.",
    317: "NO FUERCES PROCESOS.",
    318: "HOY SOSTÉN TU CENTRO.",
    319: "EL SILENCIO PROTEGE.",
    320: "TU ENERGÍA RESPONDE.",
    321: "NO CARGUES LO INNECESARIO.",
    322: "HOY BAJA EL RITMO.",
    323: "LA CALMA GUÍA.",
    324: "NO TE ADELANTES.",
    325: "HOY CONFÍA EN TU PROCESO.",
    326: "EL EQUILIBRIO SE AFINA.",
    327: "NO FUERCES ACUERDOS.",
    328: "HOY ESCUCHA MÁS.",
    329: "LA PRESENCIA BASTA.",
    330: "TU CENTRO SOSTIENE.",
    331: "NO TODO ES URGENTE.",
    332: "HOY HONRA TU CUERPO.",
    333: "LA COHERENCIA PROTEGE.",
    334: "NO CARGUES RUIDO.",
    335: "HOY ELIGE CALMA.",
    336: "EL SILENCIO ACLARA.",
    337: "NO TE DISPERSES.",
    338: "HOY VUELVE A LO ESENCIAL.",
    339: "LA CLARIDAD SE SIENTE.",
    340: "TU ENERGÍA RESPONDE.",
    341: "NO FUERCES DECISIONES.",
    342: "HOY BAJA LA EXIGENCIA.",
    343: "LA CALMA ES PODER.",
    344: "NO CARGUES TENSIONES.",
    345: "HOY CUIDA TU CENTRO.",
    346: "EL EQUILIBRIO GUÍA.",
    347: "NO REACCIONES POR HÁBITO.",
    348: "HOY CONFÍA EN TI.",
    349: "LA PRESENCIA BASTA.",
    350: "TU ENERGÍA SE ORDENA.",
    351: "NO TODO SE EXPLICA.",
    352: "HOY ESCUCHA TU INTUICIÓN.",
    353: "LA COHERENCIA SOSTIENE.",
    354: "NO FUERCES RITMOS.",
    355: "HOY ELIGE SOBRIEDAD.",
    356: "EL SILENCIO PROTEGE.",
    357: "NO TE DISPERSES.",
    358: "HOY VUELVE A TU EJE.",
    359: "LA CLARIDAD SE SIENTE.",
    360: "TU CENTRO RESPONDE.",
    361: "NO CARGUES LO INNECESARIO.",
    362: "HOY BAJA EL RUIDO.",
    363: "LA CALMA GUÍA.",
    364: "NO TE ADELANTES.",
    365: "CIERRA EL AÑO EN COHERENCIA Y VERDAD."
}



def energia_del_dia(hoy: date) -> str:
    return ENERGIA_DIA_365.get(dia_del_ano(hoy), "Hoy: respira, ordena y elige con amor.")
COMPATIBILIDAD_EXPRES = {
 
    1: (
        "Esta relación se construye desde la iniciativa y la fuerza personal.\n"
        "Ambos sienten el impulso de avanzar y liderar.\n"
        "Existe admiración mutua cuando se respetan los espacios.\n"
        "El reto aparece cuando ninguno quiere ceder.\n"
        "La relación pide reconocer al otro sin competir.\n"
        "El amor crece cuando hay apoyo y no imposición.\n"
        "Es un vínculo que necesita objetivos compartidos.\n"
        "La admiración sostiene el deseo.\n"
        "La independencia es una base, no una amenaza.\n"
        "Cuando se acompañan, avanzan con más claridad.\n"
        "La relación florece con respeto.\n"
        "El orgullo debe transformarse en cooperación.\n"
        "Ambos aprenden a liderar juntos.\n"
        "El amor se fortalece con reconocimiento.\n"
        "La unión se consolida cuando hay propósito común."
    ),

    2: (
        "Esta relación se basa en la sensibilidad y el acompañamiento emocional.\n"
        "Existe una fuerte necesidad de cercanía.\n"
        "Ambos perciben profundamente al otro.\n"
        "La relación busca cooperación y apoyo mutuo.\n"
        "El riesgo es perder la individualidad.\n"
        "El amor crece cuando hay equilibrio entre dar y recibir.\n"
        "Es un vínculo que se nutre del cuidado.\n"
        "La ternura es un lenguaje central.\n"
        "La relación se resiente si uno se anula.\n"
        "La clave está en apoyarse sin depender.\n"
        "El vínculo se fortalece con diálogo emocional.\n"
        "La unión es suave, pero profunda.\n"
        "Ambos aprenden a sostenerse.\n"
        "El amor se expresa en gestos pequeños.\n"
        "La relación prospera con armonía consciente."
    ),

    3: (
        "Esta relación se construye a través de la comunicación consciente.\n"
        "El vínculo necesita palabra, expresión y diálogo constante.\n"
        "Ambos se estimulan mental y emocionalmente.\n"
        "La creatividad es un puente de unión.\n"
        "Cuando callan lo que sienten, surge distancia.\n"
        "El cuerpo de la relación es la conversación.\n"
        "Existe potencial para alegría compartida.\n"
        "También puede aparecer dispersión emocional.\n"
        "El vínculo mejora al expresar necesidades reales.\n"
        "No se trata de hablar más, sino de hablar con verdad.\n"
        "La relación pide escucha activa.\n"
        "Cuando se comunican desde el corazón, crecen.\n"
        "El humor sana tensiones.\n"
        "La relación florece con autenticidad.\n"
        "El amor se sostiene en la palabra clara."
    ),

    4: (
        "Esta relación busca estabilidad, orden y compromiso.\n"
        "Ambos necesitan seguridad emocional.\n"
        "El vínculo se construye paso a paso.\n"
        "La constancia es una base importante.\n"
        "El riesgo es caer en rigidez.\n"
        "La relación crece cuando hay flexibilidad.\n"
        "El amor se expresa en hechos concretos.\n"
        "Ambos valoran la lealtad.\n"
        "El vínculo se fortalece con acuerdos claros.\n"
        "La rutina puede ser sostén o desgaste.\n"
        "La clave es renovar sin destruir.\n"
        "El compromiso une profundamente.\n"
        "La relación se vuelve sólida con confianza.\n"
        "Ambos aprenden a sostenerse en el tiempo.\n"
        "El amor se consolida con coherencia."
    ),

    5: (
        "Esta relación está marcada por el cambio y la libertad.\n"
        "Ambos necesitan movimiento.\n"
        "El vínculo se alimenta de experiencias compartidas.\n"
        "La rutina debilita la conexión.\n"
        "El reto es sostener continuidad emocional.\n"
        "La relación florece con acuerdos claros.\n"
        "Existe curiosidad mutua.\n"
        "La atracción se renueva con novedad.\n"
        "El riesgo es la inestabilidad.\n"
        "La libertad necesita responsabilidad.\n"
        "El amor crece cuando hay confianza.\n"
        "Ambos aprenden a elegir conscientemente.\n"
        "La relación se expande con flexibilidad.\n"
        "El vínculo se fortalece con honestidad.\n"
        "El amor se sostiene con compromiso libre."
    ),

    6: (
        "Esta relación se basa en el cuidado y la protección.\n"
        "Existe una fuerte energía de hogar.\n"
        "Ambos buscan contención emocional.\n"
        "El amor se expresa en responsabilidad afectiva.\n"
        "El riesgo es sobrecargarse.\n"
        "La relación necesita equilibrio.\n"
        "Cuidar no es controlar.\n"
        "El vínculo se fortalece con ternura.\n"
        "La familia y el entorno pesan.\n"
        "El amor madura con límites sanos.\n"
        "Ambos aprenden a dar sin agotarse.\n"
        "La relación florece con reciprocidad.\n"
        "El compromiso es profundo.\n"
        "La unión se nutre del respeto.\n"
        "El amor se sostiene con cuidado consciente."
    ),

    7: (
        "Esta relación es introspectiva y profunda.\n"
        "Existe conexión espiritual.\n"
        "Ambos necesitan espacios personales.\n"
        "El silencio también comunica.\n"
        "El riesgo es el aislamiento.\n"
        "La relación crece con comprensión.\n"
        "No todo se expresa con palabras.\n"
        "El vínculo se fortalece con confianza.\n"
        "La conexión es sutil pero intensa.\n"
        "El amor pide paciencia.\n"
        "Ambos aprenden a respetar procesos internos.\n"
        "La unión se afina con conciencia.\n"
        "El vínculo se profundiza con honestidad.\n"
        "La relación madura lentamente.\n"
        "El amor se sostiene desde la verdad interior."
    ),

    8: (
        "Esta relación es intensa y orientada a objetivos.\n"
        "Existe ambición compartida.\n"
        "Ambos buscan crecer.\n"
        "El poder puede unir o separar.\n"
        "El reto es evitar luchas de control.\n"
        "La relación florece con respeto mutuo.\n"
        "El amor se fortalece con equilibrio emocional.\n"
        "La unión pide sensibilidad.\n"
        "El éxito compartido une.\n"
        "La relación se debilita sin empatía.\n"
        "Ambos aprenden a liderar juntos.\n"
        "El vínculo madura con conciencia.\n"
        "El amor necesita humanidad.\n"
        "La relación se equilibra con humildad.\n"
        "El vínculo prospera con coherencia."
    ),

    9: (
        "Esta relación es profundamente transformadora.\n"
        "Remueve memorias emocionales.\n"
        "Existe aprendizaje mutuo.\n"
        "El vínculo invita a sanar.\n"
        "El reto es soltar el pasado.\n"
        "La relación pide compasión.\n"
        "El amor crece con perdón.\n"
        "No es una relación ligera.\n"
        "La unión cierra ciclos.\n"
        "Ambos evolucionan.\n"
        "El vínculo se profundiza con aceptación.\n"
        "La relación libera cargas emocionales.\n"
        "El amor se vuelve consciente.\n"
        "El vínculo transforma a ambos.\n"
        "La unión deja huella."
    ),

    11: (
        "Esta relación es altamente sensible e intuitiva.\n"
        "Existe conexión energética fuerte.\n"
        "Ambos perciben emociones profundas.\n"
        "El vínculo es inspirador.\n"
        "El reto es sostener lo práctico.\n"
        "La relación florece con coherencia.\n"
        "La intuición guía el vínculo.\n"
        "El amor es sutil.\n"
        "La relación puede ser intensa.\n"
        "Ambos deben cuidarse emocionalmente.\n"
        "El vínculo pide equilibrio.\n"
        "La unión inspira crecimiento.\n"
        "La relación se fortalece con verdad.\n"
        "El amor es profundo.\n"
        "La conexión es espiritual."
    ),

    22: (
        "Esta relación tiene propósito y visión compartida.\n"
        "Ambos sienten misión conjunta.\n"
        "El vínculo busca construir algo duradero.\n"
        "El reto es no cargar demasiado.\n"
        "La relación pide organización.\n"
        "El amor crece con estructura.\n"
        "La unión se fortalece con metas claras.\n"
        "El compromiso es profundo.\n"
        "Ambos se apoyan.\n"
        "El vínculo se consolida con paciencia.\n"
        "La relación madura con esfuerzo consciente.\n"
        "El amor se sostiene en hechos.\n"
        "La unión deja legado.\n"
        "El vínculo se fortalece con coherencia.\n"
        "La relación construye futuro."
    ),

    33: (
        "Esta relación es de amor profundo y servicio mutuo.\n"
        "Existe compasión intensa.\n"
        "Ambos sienten responsabilidad emocional.\n"
        "El amor es incondicional.\n"
        "El reto es no sacrificarse en exceso.\n"
        "La relación pide límites sanos.\n"
        "El vínculo sana.\n"
        "La unión es transformadora.\n"
        "El amor es generoso.\n"
        "Ambos deben cuidarse.\n"
        "La relación florece con equilibrio.\n"
        "El vínculo se fortalece con conciencia.\n"
        "La unión eleva.\n"
        "El amor es profundo.\n"
        "La relación es sanadora."
    ),
}
def compatibilidad_numero(fecha_a: date, fecha_b: date) -> int:
    return reducir_numero(
        (fecha_a.day + fecha_a.month + fecha_a.year) +
        (fecha_b.day + fecha_b.month + fecha_b.year)
    )

def compatibilidad_express_texto(n: int) -> str:
    return COMPATIBILIDAD_EXPRES.get(int(n), "Compatibilidad express no disponible.")


# =====================================================
# TEXTOS PROFUNDOS (10–12 líneas aprox)
# Basados en tu Año Personal (ap) y modulados por mp/sp/dp
# =====================================================
NUM_RASGOS = {
    1: ("iniciativa", "afirmación", "dirección"),
    2: ("sensibilidad", "cooperación", "armonía"),
    3: ("expresión", "creatividad", "comunicación"),
    4: ("estructura", "disciplina", "constancia"),
    5: ("cambio", "libertad", "movimiento"),
    6: ("cuidado", "responsabilidad", "vínculos"),
    7: ("introspección", "análisis", "intuición"),
    8: ("logro", "poder personal", "materialización"),
    9: ("cierre", "compasión", "integración"),
    11: ("inspiración", "intuición elevada", "visión"),
    22: ("construcción", "visión práctica", "impacto"),
    33: ("amor consciente", "servicio", "sabiduría emocional"),
}

def parrafo_premium_categoria(ap: int, mp: int, sp: int, dp: int, categoria: str) -> str:
    a, b, c = NUM_RASGOS.get(ap, ("equilibrio", "conciencia", "claridad"))

    base = (
        f"En {categoria}, tu ciclo se ordena desde la vibración {ap}: un núcleo de {a} que marca el ritmo principal del año. "
        f"Esto no es teoría: es una energía que se nota en decisiones, personas que aparecen, límites que se piden y oportunidades que solo se abren cuando eliges con presencia."
    )
    detalle = (
        f"Tu Mes Personal {mp} ajusta el clima emocional y práctico de este momento, y tu Semana Personal {sp} revela el tema inmediato que está ‘pidiendo voz’. "
        f"Hoy, con Día Personal {dp}, la vida te muestra en pequeño lo que debes practicar en grande: coherencia, enfoque y verdad."
    )
    guia = (
        f"La llave está en refinar tu {b} y tu {c}: no reaccionar, sino decidir. "
        f"Si {categoria.lower()} se siente tenso, no es castigo: es señal de reorden. "
        f"El movimiento correcto es simple: un límite sano, una conversación clara o un hábito que te sostenga. "
        f"Cuando actúas alineada con tu vibración, el resultado se siente: menos desgaste, más paz, y una sensación real de avance."
    )
    return f"{base}\n\n{detalle}\n\n{guia}"

# =====================================================
# PINÁCULO + ARCANO (micro)
# =====================================================
def pinaculo_micro(pin: dict) -> str:
    b1, b2, b3 = pin["base"]
    m1, m2 = pin["medio"]
    cima = pin["cima"]
    return (
        f"Tu pináculo muestra cómo se ordena tu crecimiento por etapas: la base ({b1}, {b2}, {b3}) describe aprendizajes que te forman; "
        f"el nivel medio ({m1}, {m2}) revela el punto donde se afina tu carácter; y la cima ({cima}) marca la síntesis de tu fuerza interna. "
        "Úsalo como brújula: cuando alineas hábitos y decisiones con esta estructura, avanzas con más dirección y menos desgaste."
    )

ARCANOS_RESUMIDOS = {
    1: "Inicio consciente: una decisión clara abre camino.",
    2: "Escucha interior: la respuesta se forma desde adentro.",
    3: "Creatividad: nutre lo que está creciendo.",
    4: "Orden: estructura y límites te devuelven estabilidad.",
    5: "Aprendizaje: elige desde valores, no desde presión.",
    6: "Elección: coherencia entre deseo y verdad.",
    7: "Dirección: enfoque y disciplina para avanzar.",
    8: "Equilibrio: ordena lo pendiente con honestidad.",
    9: "Introspección: comprender primero mejora tu decisión.",
    10: "Cambio: adaptarte te abre oportunidades.",
    11: "Fortaleza: calma interna por encima de la reacción.",
    12: "Nueva mirada: cambia el ángulo y aparece la salida.",
    13: "Transformación: cerrar a tiempo libera espacio.",
    14: "Armonía: ajusta extremos y cuida tu ritmo.",
    15: "Conciencia: reconoce lo que ata para recuperar poder.",
    16: "Ruptura: cae lo falso para reconstruir con verdad.",
    17: "Esperanza: guía interna y visión más amable.",
    18: "Sensibilidad: cuida emociones, evita decidir por miedo.",
    19: "Claridad: vitalidad y confianza para avanzar.",
    20: "Renacer: cierre consciente y elección con propósito.",
    21: "Integración: culminación y preparación del siguiente ciclo.",
    22: "Apertura: comienza con confianza y presencia.",
}

def arcano_micro(arc: int) -> str:
    return ARCANOS_RESUMIDOS.get(arc, "Mensaje no disponible.")

# =====================================================
# PDF helper
# =====================================================
def build_pdf_bytes(titulo: str, secciones: list[tuple[str, str]]) -> bytes:
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=LETTER)
    _, height = LETTER
    x = 50
    y = height - 60

    c.setFont("Helvetica-Bold", 16)
    c.drawString(x, y, titulo)
    y -= 22

    c.setFont("Helvetica", 10)
    c.drawString(x, y, f"{BRAND} · Generado automáticamente")
    y -= 18

    def draw_paragraph(text: str, y: int):
        c.setFont("Helvetica", 11)
        lines = []
        for para in str(text).split("\n"):
            para = para.strip()
            if not para:
                lines.append("")
                continue
            lines.extend(textwrap.wrap(para, width=95))
            lines.append("")
        for ln in lines:
            if y < 90:
                c.showPage()
                y = height - 60
            c.drawString(x, y, ln)
            y -= 14
        return y

    for head, body in secciones:
        if y < 120:
            c.showPage()
            y = height - 60
        c.setFont("Helvetica-Bold", 13)
        c.drawString(x, y, head)
        y -= 18
        y = draw_paragraph(body, y)
        y -= 6

    c.save()
    buffer.seek(0)
    return buffer.read()

# =====================================================
# CLAVE (estable, reutilizable infinitamente)
# =====================================================
def normalizar_clave_nombre(txt: str) -> str:
    txt = unicodedata.normalize("NFD", str(txt))
    txt = "".join(c for c in txt if unicodedata.category(c) != "Mn")
    txt = re.sub(r"[^A-Za-z\s]", " ", txt)
    txt = re.sub(r"\s+", " ", txt).strip().upper()
    return txt

def generar_clave_unica(nombre_completo: str, fecha_nac: date) -> str:
    nombre_normalizado = normalizar_clave_nombre(nombre_completo)
    payload = f"{nombre_normalizado}|{fecha_nac.isoformat()}".encode("utf-8")
    digest = hmac.new(APP_SECRET.encode("utf-8"), payload, hashlib.sha256).hexdigest().upper()
    core = digest[:16]
    return f"EM-{core[:4]}-{core[4:8]}-{core[8:12]}-{core[12:16]}"

# =====================================================
# TEXTO INTRO
# =====================================================
st.markdown("""
Esta lectura no es una predicción ni una promesa externa.  
Es una orientación energética consciente, basada en la vibración que se activa a partir de tu fecha de nacimiento y tu nombre.  
Cada nombre refleja una frecuencia, y cada frecuencia describe una forma de transitar la vida en este momento.

Aquí no buscamos decirte qué va a pasar, sino ayudarte a comprender qué energía está disponible para ti ahora, cómo se manifiesta internamente y qué tipo de decisiones se alinean mejor con tu proceso actual.  
La numerología, cuando se usa con consciencia, no limita: ordena, revela y enfoca.

✨ Esta lectura no te quita responsabilidad: te la devuelve.  
Tómala como una brújula, no como un destino.
""")


hoy = date.today()
dia_del_ano = hoy.timetuple().tm_yday  # 1..365

mensaje_universal = ENERGIA_DIA_365.get(
    dia_del_ano,
    "Hoy es un día para observar, integrar y no forzar."
)

st.markdown("### 😇 Mensaje universal del día")
st.write(mensaje_universal)

# =====================================================
# INPUTS
# =====================================================
col1, col2 = st.columns(2)
with col1:
    fecha_nac = st.date_input(
        "Fecha de nacimiento",
        min_value=date(1940, 1, 1),
        max_value=date(2040, 12, 31),
        value=date(1990, 1, 1),
    )
with col2:
    nombre = st.text_input(
        "Nombre completo (máx. 40 caracteres)",
        max_chars=40,
        value="",
        placeholder="Ej: Eugenia Mystikos"
    )
st.markdown("### 💞 Compatibilidad (opcional)")
activar_compat_express = st.checkbox("Activar compatibilidad express", value=False)

fecha_pareja_express = st.date_input(
    "Fecha de nacimiento de la pareja",
    min_value=date(1936, 1, 1),
    max_value=date(2036, 12, 31),
    value=date(2000, 1, 1),
    disabled=not activar_compat_express
)
calcular = st.button("✨ Ver mi lectura ahora")
hoy = date.today()

# =====================================================
# CÁLCULOS
# =====================================================
es = esencia(fecha_nac)
mis = sendero_vida(fecha_nac)
vp = vida_pasada(fecha_nac)

ap = ano_personal(fecha_nac, hoy.year)
mp = mes_personal(ap, hoy.month)
sp = semana_personal(mp, hoy.isocalendar()[1])
dp = dia_personal(mp, hoy.day)

arc = arcano_semanal()
pin = pinaculo_piramide(fecha_nac)
num_nombre = numero_nombre(nombre) if nombre.strip() else 0



# =====================================================
# MOSTRAR RESUMIDA (GRATIS) SOLO AL PRESIONAR BOTÓN
# =====================================================
if calcular:
    incrementar_contador()

    with st.container():
        st.markdown("### ✨ Tu lectura resumida")

        # AÑO PERSONAL PRIMERO (más fuerte)
        st.write(f"🔥 Vibración de tu Año Personal ({hoy.year}) — Número {ap}")
        st.write(lectura_resumida(ap))
        st.markdown(
            "Este año funciona como tu campo de experiencia principal: ordena decisiones, cierres y oportunidades. "
            "Si actúas alineada con esta vibración, la vida se vuelve más clara: menos fricción, más coherencia, y un rumbo interno más firme."
        )

        st.write(f"Mi esencia — Número {es}")
        st.write(lectura_resumida(es))

        st.write(f"Mi nombre completo — Número {num_nombre if num_nombre else '—'}")
        if num_nombre:
            st.write(lectura_resumida(num_nombre))
        else:
            st.info("Escribe tu nombre completo para ver la energía de tu nombre.")

        st.write(f"Mi misión — Número {mis}")
        st.write(lectura_resumida(mis))

        st.write(f"Mi energía de hoy — Número {dp}")
        st.write(lectura_resumida(dp))

       

        if activar_compat_express:
            comp_ex = compatibilidad_numero(fecha_nac, fecha_pareja_express)
            st.markdown(f"### 💞 Compatibilidad Express · Número {comp_ex}")
            st.write(compatibilidad_express_texto(comp_ex))


        # ✅ AQUÍ VAN LOS 4 BLOQUES CORTOS GRATIS (lo que me pediste)
        st.markdown("#### 💡 Pronóstico clave")
        st.write(frase_categoria(FRASES_AMOR, ap))
        st.write(frase_categoria(FRASES_DINERO, ap))
        st.write(frase_categoria(FRASES_EMOCIONAL, ap))
        st.write(frase_categoria(FRASES_PROTECCION, ap))

        st.write("Mi pináculo (pirámide completa)")
        st.write(f"Base: {pin['base']} | Medio: {pin['medio']} | Cima: {pin['cima']}")
        st.write(pinaculo_micro(pin))

        st.write(f"Arcano semanal — Número {arc}")
        st.write(arcano_micro(arc))

    # PDF Resumido
    pdf_resumido = build_pdf_bytes(
        f"{APP_TITLE} · Versión Resumida · {BRAND}",
        [
            ("Datos", f"Nombre: {nombre or '—'}\nFecha de nacimiento: {fecha_nac}\nGenerado: {hoy}"),
            ("Año personal", f"Número {ap}\n\n{lectura_resumida(ap)}\n\n"
                            "Este año funciona como tu campo de experiencia principal: ordena decisiones, cierres y oportunidades. "
                            "Si actúas alineada con esta vibración, la vida se vuelve más clara: menos fricción, más coherencia."),
            ("Mi esencia", f"Número {es}\n\n{lectura_resumida(es)}"),
            ("Mi nombre completo", f"Número {num_nombre if num_nombre else '—'}\n\n{lectura_resumida(num_nombre) if num_nombre else 'Escribe tu nombre completo para ver esta sección.'}"),
            ("Mi misión", f"Número {mis}\n\n{lectura_resumida(mis)}"),
            ("Mi energía de hoy", f"Número {dp}\n\n{lectura_resumida(dp)}"),
            ("Pronóstico clave (gratis)",
             f"{frase_categoria(FRASES_AMOR, ap)}\n{frase_categoria(FRASES_DINERO, ap)}\n{frase_categoria(FRASES_EMOCIONAL, ap)}\n{frase_categoria(FRASES_PROTECCION, ap)}"),
            ("Mi pináculo (pirámide completa)", f"Base: {pin['base']} | Medio: {pin['medio']} | Cima: {pin['cima']}\n\n{pinaculo_micro(pin)}"),
            ("Arcano semanal", f"Número {arc}\n\n{arcano_micro(arc)}"),
        ]
    )

    st.download_button(
        "⬇️ Descargar PDF (Versión Resumida)",
        data=pdf_resumido,
        file_name=f"Lectura_Numerologica_Resumida_{BRAND}.pdf",
        mime="application/pdf",
    )
else:
    st.caption("Tip: completa tu nombre y fecha, luego toca el botón para ver tu lectura.")

# =====================================================
# PANEL ADMIN (OCULTO POR PIN) - SOLO AQUÍ SE VE CONTADOR Y GENERADOR
# =====================================================
if ADMIN_PIN:
    with st.expander("🔐 Eugenia Mystikos (Admin)", expanded=False):
        pin_ingresado = st.text_input("PIN de administración", type="password")
        if pin_ingresado:
            if pin_ingresado == ADMIN_PIN:
                st.success("Acceso concedido ✅")
                st.info(f"📊 Uso interno · Total activaciones resumida: {leer_contador()}")
                if nombre.strip():
                    st.caption("Clave del cliente (según nombre+fecha actuales):")
                    st.code(generar_clave_unica(nombre, fecha_nac), language="text")
            else:
                st.error("PIN incorrecto")


# =========================================================
# 🔐 VERSIÓN COMPLETA (PAGO) - BLOQUEO POR CLAVE + NOMBRE + FECHA
# =========================================================

st.markdown("---")
st.markdown("## 🔐 Versión Completa (Premium + PDF personalizado)")
st.write("Desbloquea tu lectura completa con tu clave personal.")

colv1, colv2 = st.columns(2)

with colv1:
    nombre_compra = st.text_input(
        "Nombre (exactamente como en tu compra)",
        key="nombre_compra",
        max_chars=40,
        placeholder="Ej: Eugenia Mystikos"
    )

with colv2:
    fecha_compra = st.date_input(
        "Fecha de nacimiento (como en tu compra)",
        key="fecha_compra",
        min_value=date(1940, 1, 1),
        max_value=date(2040, 12, 31),
        value=date(1990, 1, 1),
    )

clave_ingresada = st.text_input(
    "Introduce tu clave personal",
    type="password"
).strip().upper()

# 👉 BOTÓN CLAVE (ESTO ES LO QUE FALTABA)
confirmar_datos = st.button("🔓 Confirmar datos y desbloquear")

# =========================================================
# VALIDACIÓN (SOLO SE EJECUTA AL PRESIONAR EL BOTÓN)
# =========================================================

if confirmar_datos:

    if not nombre_compra.strip():
        st.warning("Escribe tu nombre tal como aparece en tu compra.")
        st.stop()

    if not fecha_compra:
        st.warning("Debes indicar la fecha de nacimiento usada en tu compra.")
        st.stop()

    if not clave_ingresada:
        st.warning("Debes introducir tu clave personal.")
        st.stop()

    clave_esperada = generar_clave_unica(nombre_compra, fecha_compra)

    if clave_ingresada != clave_esperada:
        st.error("Clave inválida. Verifica que tu nombre y fecha estén EXACTAMENTE como en tu compra.")
        st.stop()

       

# #####################################################
# =========================================================
# 📘 MOTOR PREMIUM (PYTHON CALCULA TODO + DICCIONARIO EXCEL + PDF BONITO)
# =========================================================
########################################################################


# =========================
# CONFIG
# =========================
MAESTROS = {11, 22, 33, 44}

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DICC_PATH = os.path.join(BASE_DIR, "Diccionario.xlsx")

# Paleta Eugenia Mística
COLOR_ROJO_MISTICO = "#7A1E3A"
COLOR_DORADO = "#9C7A3F"
COLOR_TEXTO = "#2E2E2E"
COLOR_GRIS = "#666666"

# Año actual (para año personal / cuatrimestres / etc.)
HOY = date.today()
ANO_ACTUAL = HOY.year
def personalizar_texto(texto: str, nombre: str) -> str:
    if not texto:
        return texto

    nombre = nombre.strip()

    reglas = {
        # Referencias impersonales → personales
        "Las personas nacidas en": f"{nombre}, al vibrar en",
        "Las personas que nacen en": f"{nombre}, al vibrar en",
        "Estas personas": "Tú",
        "Estas almas": "Tu alma",
        "Estos individuos": "Tú",
        "Ellos": "Tú",
        "Ellas": "Tú",

        # Vida / camino
        "Su vida": "Tu vida",
        "Su camino": "Tu camino",
        "Su misión": "Tu misión",
        "Su energía": "Tu energía",
        "Su vibración": "Tu vibración",

        # Conducta
        "tienden a": "tiendes a",
        "suelen": "sueles",
        "pueden": "puedes",
        "deben": "debes",

        # Lenguaje distante → cercano
        "Se observa que": "La vida te muestra que",
        "Esto indica que": "Esto te indica que",
        "Esto sugiere que": "Esto te sugiere que",
        "Es importante que": "Es importante para ti que",
    }

    for origen, destino in reglas.items():
        texto = texto.replace(origen, destino)

    return texto

# =========================
# UTILIDADES TEXTO / NOMBRE
# =========================
def _norm_txt(s: str) -> str:
    s = (s or "").strip()
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = s.replace("ñ", "n").replace("Ñ", "N")
    s = re.sub(r"\s+", " ", s)
    return s

def _solo_letras(s: str) -> str:
    s = _norm_txt(s).upper()
    s = re.sub(r"[^A-Z ]", "", s)
    return s

def separar_nombre_apellido(full_name: str):
    """
    Heurística:
    - Si hay 4+ tokens: 2 primeros = nombre(s), resto = apellido(s)
    - Si hay 3 tokens: 1 primero = nombre, resto = apellidos
    - Si hay 2 tokens: 1 primero = nombre, 1 segundo = apellido
    - Si hay 1 token: todo nombre
    """
    tokens = _solo_letras(full_name).split()
    if len(tokens) >= 4:
        nombre = " ".join(tokens[:2])
        apellido = " ".join(tokens[2:])
    elif len(tokens) == 3:
        nombre = tokens[0]
        apellido = " ".join(tokens[1:])
    elif len(tokens) == 2:
        nombre = tokens[0]
        apellido = tokens[1]
    else:
        nombre = " ".join(tokens) if tokens else ""
        apellido = ""
    return nombre, apellido


# =========================
# NUMEROLOGÍA BÁSICA
# =========================
def suma_digitos(n: int) -> int:
    return sum(int(d) for d in str(abs(int(n))))

def reducir_con_maestros(n: int) -> int:
    """
    Reduce a 1-9, pero detiene en 11/22/33/44.
    """
    n = abs(int(n))
    while n > 9 and n not in MAESTROS:
        n = suma_digitos(n)
    return n

def reducir_estricto_1a9(n: int) -> int:
    """
    Reduce SIEMPRE hasta 1-9 (ignora maestros).
    (Esto aplica a Animal Espiritual, Tarot repetidos, Salud/Espíritu, etc. según tu nota.)
    """
    n = abs(int(n))
    while n > 9:
        n = suma_digitos(n)
    return n

def reducir_excepcion_10_11(n: int) -> int:
    """
    Para Don Divino: reduce a 1-9 salvo si cae en 10 o 11.
    """
    n = abs(int(n))
    while n > 11 and n not in {10, 11}:
        n = suma_digitos(n)
    return n

def reducir_a_dos_digitos(n: int) -> int:
    """
    Reduce por suma de dígitos hasta quedar en 1..99.
    """
    n = abs(int(n))
    while n >= 100:
        n = suma_digitos(n)
    return n

def regla_tarot_78(n: int) -> int:
    """
    Si el resultado es < 78, se deja (puede ser 2 dígitos).
    Si no, se reduce a 1-9 (estricto).
    """
    n = abs(int(n))
    if n < 78:
        return n
    return reducir_estricto_1a9(n)

def suma_ano_en_digitos(year: int) -> int:
    return suma_digitos(year)


# =========================
# VALORES LETRAS (PITAGÓRICO)
# =========================
# 1: A J S
# 2: B K T
# 3: C L U
# 4: D M V
# 5: E N W
# 6: F O X
# 7: G P Y
# 8: H Q Z
# 9: I R
MAPA_LETRA = {}
for ch in "AJS": MAPA_LETRA[ch] = 1
for ch in "BKT": MAPA_LETRA[ch] = 2
for ch in "CLU": MAPA_LETRA[ch] = 3
for ch in "DMV": MAPA_LETRA[ch] = 4
for ch in "ENW": MAPA_LETRA[ch] = 5
for ch in "FOX": MAPA_LETRA[ch] = 6
for ch in "GPY": MAPA_LETRA[ch] = 7
for ch in "HQZ": MAPA_LETRA[ch] = 8
for ch in "IR":  MAPA_LETRA[ch] = 9

VOCALES = set("AEIOU")

def valor_letra(ch: str) -> int:
    ch = _solo_letras(ch).replace(" ", "")
    if not ch:
        return 0
    return MAPA_LETRA.get(ch[0], 0)

def suma_nombre(frase: str) -> int:
    frase = _solo_letras(frase).replace(" ", "")
    return sum(MAPA_LETRA.get(ch, 0) for ch in frase)

def suma_vocales(frase: str) -> int:
    frase = _solo_letras(frase).replace(" ", "")
    return sum(MAPA_LETRA.get(ch, 0) for ch in frase if ch in VOCALES)

def suma_consonantes(frase: str) -> int:
    frase = _solo_letras(frase).replace(" ", "")
    return sum(MAPA_LETRA.get(ch, 0) for ch in frase if ch not in VOCALES)

def contar_letras(frase: str) -> int:
    frase = _solo_letras(frase).replace(" ", "")
    return len(frase)

def primera_vocal_valor(frase: str) -> int:
    frase = _solo_letras(frase).replace(" ", "")
    for ch in frase:
        if ch in VOCALES:
            return MAPA_LETRA.get(ch, 0)
    return 0

def primera_consonante_valor(frase: str) -> int:
    frase = _solo_letras(frase).replace(" ", "")
    for ch in frase:
        if ch not in VOCALES:
            return MAPA_LETRA.get(ch, 0)
    return 0

def moda_numeros(frase: str):
    frase = _solo_letras(frase).replace(" ", "")
    vals = [MAPA_LETRA.get(ch, 0) for ch in frase if MAPA_LETRA.get(ch, 0) > 0]
    if not vals:
        return None
    c = Counter(vals)
    maxf = max(c.values())
    tops = sorted([k for k,v in c.items() if v == maxf])
    return tops[0]  # si hay empate, el menor


# =========================
# DICCIONARIO DESDE EXCEL
# (cada hoja = concepto; columnas: Numero | Titulo | Texto)
# =========================
def cargar_diccionario_excel(path: str):
    wb = load_workbook(path, data_only=True)
    dicc = {}
    sheet_map = {sh.strip().lower(): sh for sh in wb.sheetnames}

    for sh_low, sh_real in sheet_map.items():
        ws = wb[sh_real]
        # asumimos encabezado en fila 1 y datos desde fila 2:
        tabla = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            num = row[0]
            if num in (None, "", "None"):
                continue
            try:
                num_int = int(num)
            except:
                continue
            titulo = (row[1] if len(row) > 1 else "") or ""
            texto  = (row[2] if len(row) > 2 else "") or ""
            tabla[num_int] = {
                "titulo": str(titulo).strip(),
                "texto": str(texto).strip()
            }
        dicc[sh_low] = tabla

    return dicc

DICC = cargar_diccionario_excel(DICC_PATH)


# =========================
# BUSCAR TEXTO EN DICCIONARIO
# =========================
def dicc_get(concepto: str, numero: int):
    """
    Retorna dict {titulo,texto} o vacío.
    'concepto' debe coincidir con el nombre de la hoja (en minúscula).
    """
    key = (concepto or "").strip().lower()
    tabla = DICC.get(key, {})
    return tabla.get(int(numero), {"titulo": "", "texto": ""})


# =========================
# CÁLCULOS (1..60) SEGÚN TU ARCHIVO
# =========================
def calcular_todo(nombre_full: str, fecha_nac: date):
    nombre, apellido = separar_nombre_apellido(nombre_full)

    dd = fecha_nac.day
    mm = fecha_nac.month
    yy = fecha_nac.year

    dd_red_maestros = reducir_con_maestros(dd)
    mm_red_maestros = reducir_con_maestros(mm)
    yy_red_maestros = reducir_con_maestros(suma_ano_en_digitos(yy))  # año en dígitos

    # 1) Misión
    mision = reducir_con_maestros(dd)

    # 2) Sendero Natal (fecha completa)
    sendero_natal = reducir_con_maestros(dd + mm + suma_ano_en_digitos(yy))

    # 3) Animal Espiritual 1 (estricto 1-9)
    animal1 = reducir_estricto_1a9(dd + mm + suma_ano_en_digitos(yy))

    # 4) Animal Espiritual 2
    animal2_cand = reducir_estricto_1a9(dd_red_maestros)  # “suma del día reducida a un dígito”
    animal2 = None if animal2_cand == animal1 else animal2_cand

    # 5) Día de nacimiento sin reducir
    dia_nac = dd

    # 6) Primer Tarot (estricto 1-9)
    tarot1 = reducir_estricto_1a9(dd + mm + suma_ano_en_digitos(yy))

    # 7) Segundo Tarot
    tarot2_cand = reducir_estricto_1a9(dd + mm + suma_ano_en_digitos(yy))
    tarot2 = None if tarot2_cand == tarot1 else tarot2_cand

    # 8) Salud y Espíritu 1
    salud1 = reducir_estricto_1a9(dd + mm + suma_ano_en_digitos(yy))

    # 9) Salud y Espíritu 2
    salud2_cand = reducir_estricto_1a9(dd + mm + suma_ano_en_digitos(yy))
    salud2 = None if salud2_cand == salud1 else salud2_cand

    # 10) Arquetipo de Amante
    amante = reducir_estricto_1a9(dd + mm + suma_ano_en_digitos(yy))

    # 11) Vincular
    vincular = reducir_estricto_1a9(dd + mm + suma_ano_en_digitos(yy))

    # 12) Lección de Vida
    leccion_vida = reducir_estricto_1a9(dd + mm + suma_ano_en_digitos(yy))

    # 13) Primer Desafío = |dia reducido - mes reducido|
    primer_desafio = abs(reducir_con_maestros(dd) - reducir_con_maestros(mm))

    # 14) Segundo Desafío = |dia reducido - año reducido|
    segundo_desafio = abs(reducir_con_maestros(dd) - reducir_con_maestros(suma_ano_en_digitos(yy)))

    # 15) Don Divino = suma dos últimas cifras del año, reduce salvo 10/11
    ult2 = yy % 100
    don_divino = reducir_excepcion_10_11(suma_digitos(ult2))

    # 16) Nro de Raíz = si (dia+mes+año) < 10 => no posee
    total_raiz = dd + mm + yy
    nro_raiz = None if total_raiz < 10 else reducir_estricto_1a9(total_raiz)

    # 17) Esencia = vocales(nombre)+vocales(apellido) reduce con maestros
    esencia = reducir_con_maestros(suma_vocales(nombre) + suma_vocales(apellido))

    # 18) Imagen = consonantes(nombre)+consonantes(apellido) reduce con maestros
    imagen = reducir_con_maestros(suma_consonantes(nombre) + suma_consonantes(apellido))

    # 19) Destino = suma(nombre)+suma(apellido) reduce con maestros
    destino = reducir_con_maestros(suma_nombre(nombre) + suma_nombre(apellido))

    # 20) Nro Letras Nombre (sin espacios)
    nro_letras = contar_letras(nombre + apellido)

    # 21..25 años importantes
    anio_imp_1 = nro_letras * 1
    anio_imp_2 = nro_letras * 2
    anio_imp_3 = nro_letras * 3
    anio_imp_4 = nro_letras * 4
    anio_imp_5 = nro_letras * 5

    # 26) Características Vida = nro letras reducido a 1 dígito (estricto)
    caract_vida = reducir_estricto_1a9(nro_letras) if nro_letras else None

    # 27) Nro Hereditario = suma(apellido) reducido a 1 dígito (estricto)
    nro_hereditario = reducir_estricto_1a9(suma_nombre(apellido)) if apellido else None

    # 28) Talento = igual destino (según tu lista)
    talento = reducir_con_maestros(suma_nombre(nombre) + suma_nombre(apellido))

    # 29) Estado Espiritual = moda números del nombre+apellido
    estado_espiritual = moda_numeros(nombre + " " + apellido)

    # 30) Desafío Íntimo
    des_intimo = abs(primera_vocal_valor(nombre) - primera_vocal_valor(apellido))

    # 31) Desafío Realización
    des_real = abs(primera_consonante_valor(nombre) - primera_consonante_valor(apellido))

    # 32) Desafío Expresión = suma(des_intimo + des_real) reducido a 1 dígito (estricto)
    des_exp = reducir_estricto_1a9(des_intimo + des_real)

    # 33) Nro Expresión = suma(nombre+apellido) reduce con excepción 11/22 (solo)
    def reducir_solo_11_22(n: int) -> int:
        n = abs(int(n))
        while n > 9 and n not in {11, 22}:
            n = suma_digitos(n)
        return n
    nro_expresion = reducir_solo_11_22(suma_nombre(nombre) + suma_nombre(apellido))

    # 34) Potencial = Sendero Natal + Destino reducido con excepción 11/22
    potencial = reducir_solo_11_22(sendero_natal + destino)

    # 35) Años 1ra etapa = 1..(36 - suma(dia+mes+año) reducida)
    suma_fn_reducida = reducir_estricto_1a9(dd + mm + suma_ano_en_digitos(yy))
    tope_1ra = 36 - suma_fn_reducida
    if tope_1ra < 1:
        rango_1ra = "1"
    else:
        rango_1ra = f"1 - {tope_1ra}"

    # 36) Primera Etapa = (dia+mes) reducido con excepción 11/22
    primera_etapa = reducir_solo_11_22(dd + mm)

    # 37) Años 2da etapa
    ini_2da = tope_1ra + 1 if tope_1ra >= 1 else 2
    fin_2da = (tope_1ra + 10) if tope_1ra >= 1 else 11
    rango_2da = f"{ini_2da} - {fin_2da}"

    # 38) Segunda Etapa = (dia + año_dígitos) reducido con excepción 11/22
    segunda_etapa = reducir_solo_11_22(dd + suma_ano_en_digitos(yy))

    # 39) Años 3ra etapa
    ini_3ra = fin_2da + 1
    fin_3ra = fin_2da + 10
    rango_3ra = f"{ini_3ra} - {fin_3ra}"

    # 40) Tercera Etapa = (1ra + 2da) reducido con excepción 11/22
    tercera_etapa = reducir_solo_11_22(primera_etapa + segunda_etapa)

    # 41) Años 4ta etapa
    ini_4ta = fin_3ra + 1
    fin_4ta = fin_3ra + 10
    rango_4ta = f"{ini_4ta} - {fin_4ta}"

    # 42) Cuarta Etapa = (mes + año_dígitos) reducido con excepción 11/22
    cuarta_etapa = reducir_solo_11_22(mm + suma_ano_en_digitos(yy))

    # 43) Año Personal = (dia+mes+year_actual) reducido con excepción 11/22
    ano_personal = reducir_solo_11_22(dd + mm + suma_ano_en_digitos(ANO_ACTUAL))

    # 44) Dígito Edad = suma(edad + (edad-1)) reducido con excepción 11/22
    edad = ANO_ACTUAL - yy
    digito_edad = reducir_solo_11_22(edad + (edad - 1))

    # 45) Armónico = (suma año actual + suma año nac) => reduce a 2 dígitos; si <78, dejar; si no, reducir 1-9
    armonico_raw = suma_ano_en_digitos(ANO_ACTUAL) + suma_ano_en_digitos(yy)
    armonico_2d = reducir_a_dos_digitos(armonico_raw)
    armonico = armonico_2d if armonico_2d < 78 else reducir_estricto_1a9(armonico_2d)

    # 46) Tarot 1er Cuat = (suma año actual + suma año actual) - suma año nac  (regla <78)
    tarot_1c = regla_tarot_78((suma_ano_en_digitos(ANO_ACTUAL) + suma_ano_en_digitos(ANO_ACTUAL)) - suma_ano_en_digitos(yy))

    # 47) Tarot 2do Cuat = (suma año actual + dia + mes + año_nac_dígitos) (regla <78)
    tarot_2c = regla_tarot_78(suma_ano_en_digitos(ANO_ACTUAL) + dd + mm + suma_ano_en_digitos(yy))

    # 48) Tarot 3er Cuat = (suma año actual + clave personal del día y mes) (regla <78)
    # Interpretación: clave día+mes reducida con excepción 11/22
    clave_dia_mes = reducir_solo_11_22(dd + mm)
    tarot_3c = regla_tarot_78(suma_ano_en_digitos(ANO_ACTUAL) + clave_dia_mes)

    # 49..60 Meses = (año personal + k) reducida con excepción 11/22
    def mes_personal(k: int) -> int:
        return reducir_solo_11_22(ano_personal + k)

    enero = mes_personal(1)
    febrero = mes_personal(2)
    marzo = mes_personal(3)
    abril = mes_personal(4)
    mayo = mes_personal(5)
    junio = mes_personal(6)
    julio = mes_personal(7)
    agosto = mes_personal(8)
    septiembre = mes_personal(9)
    octubre = mes_personal(1)
    noviembre = mes_personal(2)
    diciembre = mes_personal(3)

    # Empaquetar resultados en el ORDEN EXACTO
    # (concepto hoja_dicc, etiqueta, valor, nota_si_no_dicc)
    items = [
        ("mision", "Misión", mision, None),
        ("sendero natal", "Sendero Natal", sendero_natal, None),
        ("animal espiritual 1", "Animal Espiritual 1", animal1, None),
        ("animal espiritual 2", "Animal Espiritual 2", animal2, "no posee segundo animal espiritual" if animal2 is None else None),
        ("dia de nacimiento", "Día de Nacimiento", dia_nac, None),
        ("primer tarot", "Primer Tarot", tarot1, None),
        ("segundo tarot", "Segundo Tarot", tarot2, "no posee segundo tarot" if tarot2 is None else None),
        ("salud y espiritu 1", "Salud y Espíritu 1", salud1, None),
        ("salud y espiritu 2", "Salud y Espíritu 2", salud2, "No existe una segunda relación entre tu espíritu y tu salud" if salud2 is None else None),
        ("arquetipo de amante", "Arquetipo de Amante", amante, None),
        ("vincular", "Vincular", vincular, None),
        ("leccion de vida", "Lección de Vida", leccion_vida, None),
        ("primer desafio", "Primer Desafío", primer_desafio, None),
        ("segundo desafio", "Segundo Desafío", segundo_desafio, None),
        ("don divino", "Don Divino", don_divino, None),
        ("nro de raiz", "Número de Raíz", nro_raiz, "No posees número de raíz" if nro_raiz is None else None),
        ("esencia", "Esencia", esencia, None),
        ("imagen", "Imagen", imagen, None),
        ("destino", "Destino", destino, None),
        ("nro letras nombre", "Nro. Letras (Nombre+Apellido)", nro_letras, None),
        ("primer año importante de tu vida", "Primer año importante", anio_imp_1, None),
        ("segundo año importante de tu vida", "Segundo año importante", anio_imp_2, None),
        ("tercer año importante de tu vida", "Tercer año importante", anio_imp_3, None),
        ("cuarto año importante de tu vida", "Cuarto año importante", anio_imp_4, None),
        ("quinto año importante de tu vida", "Quinto año importante", anio_imp_5, None),
        ("caracteristicas vida", "Características de Vida", caract_vida, None),
        ("nro hereditario", "Número Hereditario", nro_hereditario, None),
        ("talento", "Talento", talento, None),
        ("estado espiritual", "Estado Espiritual", estado_espiritual, None),
        ("desafio intimo", "Desafío Íntimo", des_intimo, None),
        ("desafio de realizacion", "Desafío de Realización", des_real, None),
        ("desafio de expresion", "Desafío de Expresión", des_exp, None),
        ("nro de expresion", "Número de Expresión", nro_expresion, None),
        ("potencial", "Potencial", potencial, None),
        ("años de la primera etapa", "Años de la Primera Etapa", rango_1ra, None),
        ("primera etapa", "Primera Etapa", primera_etapa, None),
        ("años de la segunda etapa", "Años de la Segunda Etapa", rango_2da, None),
        ("segunda etapa", "Segunda Etapa", segunda_etapa, None),
        ("años de la tercera etapa", "Años de la Tercera Etapa", rango_3ra, None),
        ("tercera etapa", "Tercera Etapa", tercera_etapa, None),
        ("años de la cuarta etapa", "Años de la Cuarta Etapa", rango_4ta, None),
        ("cuarta etapa", "Cuarta Etapa", cuarta_etapa, None),
        ("año personal", "Año Personal", ano_personal, None),
        ("digito de la edad", "Dígito de la Edad", digito_edad, None),
        ("armonico", "Armónico", armonico, None),
        ("tarot 1er cuat", "Tarot 1er Cuatrimestre", tarot_1c, None),
        ("tarot 2do cuat", "Tarot 2do Cuatrimestre", tarot_2c, None),
        ("tarot 3er cuat", "Tarot 3er Cuatrimestre", tarot_3c, None),
        ("enero", "Enero", enero, None),
        ("febrero", "Febrero", febrero, None),
        ("marzo", "Marzo", marzo, None),
        ("abril", "Abril", abril, None),
        ("mayo", "Mayo", mayo, None),
        ("junio", "Junio", junio, None),
        ("julio", "Julio", julio, None),
        ("agosto", "Agosto", agosto, None),
        ("septiembre", "Septiembre", septiembre, None),
        ("octubre", "Octubre", octubre, None),
        ("noviembre", "Noviembre", noviembre, None),
        ("diciembre", "Diciembre", diciembre, None),
    ]

    return {
        "nombre_full": _norm_txt(nombre_full),
        "nombre": nombre,
        "apellido": apellido,
        "fecha_nac": fecha_nac.strftime("%d/%m/%Y"),
        "items": items,
    }
    # =========================
# PDF BONITO (sin tablas feas, respirable)
# =========================
def build_pdf_premium(resultado: dict) -> bytes:
    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=LETTER,
        rightMargin=55,
        leftMargin=55,
        topMargin=60,
        bottomMargin=55
    )

    styles = getSampleStyleSheet()

    styles.add(ParagraphStyle(
        name="EM_TituloPortada",
        fontSize=26,
        leading=32,
        alignment=1,
        textColor=HexColor(COLOR_ROJO_MISTICO),
        spaceAfter=16
    ))

    styles.add(ParagraphStyle(
        name="EM_SubPortada",
        fontSize=13.5,
        leading=18,
        alignment=1,
        textColor=HexColor(COLOR_DORADO),
        spaceAfter=10
    ))

    styles.add(ParagraphStyle(
        name="EM_Marca",
        fontSize=10.5,
        leading=14,
        alignment=1,
        textColor=HexColor(COLOR_GRIS),
        spaceBefore=18
    ))

    styles.add(ParagraphStyle(
        name="EM_TituloSeccion",
        fontSize=15.5,
        leading=21,
        textColor=HexColor(COLOR_ROJO_MISTICO),
        spaceBefore=18,
        spaceAfter=10
    ))

    styles.add(ParagraphStyle(
        name="EM_Texto",
        fontSize=11.2,
        leading=17,
        textColor=HexColor(COLOR_TEXTO),
        spaceAfter=12
    ))

    elementos = []

    # -------------------------
    # PORTADA
    # -------------------------
    elementos.append(Spacer(1, 70))
    elementos.append(
        Paragraph("Lectura Numerológica Premium", styles["EM_TituloPortada"])
    )
    elementos.append(
        Paragraph(
            f"Informe personalizado para<br/>{resultado['nombre_full']}",
            styles["EM_SubPortada"]
        )
    )
    elementos.append(
        Paragraph(
            f"Fecha de nacimiento: {resultado['fecha_nac']}",
            styles["EM_SubPortada"]
        )
    )
    elementos.append(Spacer(1, 34))
    elementos.append(
        Paragraph(
            "Eugenia Mística · Numerología & Conciencia",
            styles["EM_Marca"]
        )
    )
    elementos.append(PageBreak())

    # -------------------------
    # CONTENIDO
    # -------------------------
    for (hoja_dicc, etiqueta, valor, nota) in resultado["items"]:

        # Título de sección
        elementos.append(
            Paragraph(etiqueta, styles["EM_TituloSeccion"])
        )

        # Resultado (misma tipografía que el texto)
        if valor is None:
            resultado_txt = "—"
        else:
            resultado_txt = str(valor)

        elementos.append(
            Paragraph(f"Resultado: {resultado_txt}", styles["EM_Texto"])
        )

        # Nota directa (si existe)
        if nota:
            elementos.append(
                Paragraph(nota, styles["EM_Texto"])
            )
            continue

        # Texto largo desde diccionario
        if isinstance(valor, int):
            info = hoja_dicc.get(valor, {})
            texto = info.get("texto", "").strip()
            texto = personalizar_texto(texto, resultado["nombre_full"])

            if texto:
                intro = f"{resultado['nombre_full']}, esta lectura se manifiesta como un espejo de tu proceso interno.\n\n"
                texto = intro + texto
                partes = [p.strip() for p in texto.split("\n") if p.strip()]
                for p in partes:
                    elementos.append(
                        Paragraph(p, styles["EM_Texto"])
                    )
            else:
                elementos.append(
                    Paragraph(
                        "No se encontró texto asociado a este resultado.",
                        styles["EM_Texto"]
                    )
                )

    doc.build(elementos)
    buffer.seek(0)
    return buffer.getvalue()
   

 # ======================================================
# ✅ DESBLOQUEO + EJECUCIÓN PREMIUM (BLOQUE FINAL ÚNICO)
# ======================================================

if confirmar_datos:
    st.session_state.premium_activo = True
    st.success("Versión completa desbloqueada ✅")

if st.session_state.premium_activo:
    resultado = calcular_todo(nombre_compra, fecha_compra)
    pdf_bytes = build_pdf_premium(resultado)

    st.download_button(
        "📄 Descargar tu Informe Premium (PDF)",
        data=pdf_bytes,
        file_name=f"Lectura_Premium_{_norm_txt(nombre_compra)}.pdf",
        mime="application/pdf",
    )
